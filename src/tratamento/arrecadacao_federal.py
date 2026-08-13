"""Etapa 3 — Tratamento da fonte RFB: Arrecadação das Receitas Federais.

O XLSX vem em formato "wide" e hierárquico: **uma aba por ano**, linhas = tributos
(o recuo do rótulo indica o nível na hierarquia) e colunas = JAN…DEZ + TOTAL.
Este módulo converte tudo para o formato **tidy/long** do projeto:

    ano_mes | tributo | rotulo | tributo_pai | nivel | tipo | valor | valor_constante

- ``valor`` é o valor corrente em R$ milhões (como publicado);
- ``valor_constante`` é o mesmo valor deflacionado pelo IPCA (SGS 433) para o
  poder de compra do **último mês da série**;
- ``tipo`` separa ``componente`` (as 15 linhas de nível 1 que somam o total geral),
  ``detalhe`` (aberturas dentro de um componente) e ``agregado`` (as linhas de
  soma do próprio XLSX) — evitando dupla contagem em participações e rankings.

Só o arquivo de 1994 em diante é tratado: o de 1970-1993 mistura quatro moedas
(Cr$, Cz$, NCz$, CR$) e traz 1970-1985 em base anual, o que não forma série
comparável (ver ``catalogo/fontes.md``). Ele é coletado apenas para rastreabilidade.

Grava o resultado em ``dados/processados/arrecadacao_federal.csv``.
"""
from __future__ import annotations

import json
import unicodedata

import openpyxl
import pandas as pd

from src import config

SLUG = "arrecadacao_federal"

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
         "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
LINHA_CABECALHO = 6  # linha "RECEITAS | JAN | FEV | ..." (1-indexada)


def _pasta_bruta(slug: str):
    pasta = config.DIR_BRUTOS / slug
    if not (pasta / "_metadados.json").exists():
        raise FileNotFoundError(
            f"Bruto ausente: {pasta}. Rode a Etapa 2 (coleta) antes do tratamento."
        )
    return pasta


def _nivel(rotulo_bruto: str) -> int:
    """Nível hierárquico a partir do recuo do rótulo (0, 2-3 e 4 espaços no XLSX)."""
    recuo = len(rotulo_bruto) - len(rotulo_bruto.lstrip())
    if recuo == 0:
        return 1
    return 2 if recuo < 4 else 3


def _chave(texto: str) -> str:
    """Normaliza rótulo para comparação (sem acento, caixa alta, espaços colapsados)."""
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return " ".join(sem_acento.upper().split())


def _ler_planilha(caminho, cfg: dict) -> pd.DataFrame:
    """Lê todas as abas anuais do XLSX e devolve as linhas tidy (ainda sem deflação)."""
    componentes = {_chave(t) for t in cfg["componentes_total"]}
    agregados = {_chave(t) for t in cfg["agregados"]}
    rotulos = {_chave(k): v for k, v in cfg["rotulos"].items()}

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    linhas = []
    for aba in wb.sheetnames:
        if not aba.isdigit():  # ignora abas com intervalos (ex.: "1970-1985")
            continue
        ano = aba
        ws = wb[aba]
        cabecalho = [c for c in next(ws.iter_rows(
            min_row=LINHA_CABECALHO, max_row=LINHA_CABECALHO, values_only=True))]
        # Posição de cada mês no cabeçalho (ignora a coluna TOTAL).
        col_mes = {}
        for i, valor in enumerate(cabecalho):
            if isinstance(valor, str) and valor.strip().upper() in MESES:
                col_mes[valor.strip().upper()] = i

        pai_por_nivel: dict[int, str] = {}
        for celulas in ws.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True):
            bruto = celulas[0]
            if not isinstance(bruto, str) or not bruto.strip():
                continue
            tributo = " ".join(bruto.split())
            nivel = _nivel(bruto)
            chave = _chave(tributo)
            pai_por_nivel[nivel] = tributo
            pai = pai_por_nivel.get(nivel - 1) if nivel > 1 else None

            if chave in agregados:
                tipo = "agregado"
            elif chave in componentes:
                tipo = "componente"
            else:
                tipo = "detalhe"

            for mes_nome, indice in col_mes.items():
                valor = celulas[indice] if indice < len(celulas) else None
                if not isinstance(valor, (int, float)):
                    continue
                linhas.append({
                    "ano_mes": f"{ano}-{MESES.index(mes_nome) + 1:02d}",
                    "tributo": tributo,
                    "rotulo": rotulos.get(chave, tributo),
                    "tributo_pai": pai,
                    "nivel": nivel,
                    "tipo": tipo,
                    "valor": float(valor),
                })
    wb.close()
    return pd.DataFrame(linhas)


def _desambiguar(tidy: pd.DataFrame) -> pd.DataFrame:
    """Qualifica com o pai os rótulos que se repetem em tributos diferentes.

    "ENTIDADES FINANCEIRAS" e "DEMAIS EMPRESAS" aparecem dentro de IRPJ, COFINS,
    PIS/PASEP e CSLL — sem qualificar, quatro séries distintas colidiriam.
    """
    pais_por_nome = tidy.groupby("tributo")["tributo_pai"].nunique(dropna=False)
    ambiguos = set(pais_por_nome[pais_por_nome > 1].index)
    if not ambiguos:
        return tidy

    duplicado = tidy["tributo"].isin(ambiguos) & tidy["tributo_pai"].notna()
    pai_rotulo = tidy.drop_duplicates("tributo").set_index("tributo")["rotulo"]
    tidy.loc[duplicado, "rotulo"] = (
        tidy.loc[duplicado, "tributo_pai"].map(pai_rotulo).fillna(
            tidy.loc[duplicado, "tributo_pai"])
        + " · " + tidy.loc[duplicado, "rotulo"]
    )
    tidy.loc[duplicado, "tributo"] = (
        tidy.loc[duplicado, "tributo_pai"] + " · " + tidy.loc[duplicado, "tributo"]
    )
    print(f"[tratamento] {len(ambiguos)} rótulos ambíguos qualificados pelo tributo pai.")
    return tidy


def _indice_ipca(pasta) -> pd.Series | None:
    """Constrói o número-índice do IPCA (base 1 no 1º mês) a partir da variação %."""
    caminho = pasta / "ipca_sgs433.json"
    if not caminho.exists():
        print("[tratamento] aviso: IPCA ausente — valores constantes não serão gerados.")
        return None
    registros = json.loads(caminho.read_text(encoding="utf-8"))
    serie = pd.DataFrame(registros)
    serie["ano_mes"] = serie["data"].str.slice(6, 10) + "-" + serie["data"].str.slice(3, 5)
    serie["variacao"] = pd.to_numeric(serie["valor"], errors="coerce") / 100
    serie = serie.dropna(subset=["variacao"]).sort_values("ano_mes")
    indice = (1 + serie["variacao"]).cumprod()
    indice.index = serie["ano_mes"].values
    return indice


def _deflacionar(tidy: pd.DataFrame, indice: pd.Series | None) -> pd.DataFrame:
    """Acrescenta ``valor_constante`` a preços do último mês coberto pela série."""
    if indice is None:
        tidy["valor_constante"] = pd.NA
        return tidy

    mes_base = max(m for m in tidy["ano_mes"].unique() if m in indice.index)
    fator = indice.loc[mes_base] / indice
    tidy["valor_constante"] = tidy["valor"] * tidy["ano_mes"].map(fator)
    faltando = tidy["valor_constante"].isna().sum()
    if faltando:
        print(f"[tratamento] aviso: {faltando} linhas sem IPCA correspondente.")
    print(f"[tratamento] valores constantes a preços de {mes_base} (IPCA/SGS 433)")
    return tidy


def _conferir_total(tidy: pd.DataFrame, cfg: dict) -> None:
    """Verifica se os componentes somam o TOTAL GERAL publicado (tolerância 0,1%)."""
    total_linha = _chave(cfg["linha_total"])
    publicado = tidy[tidy["tributo"].map(_chave) == total_linha].set_index("ano_mes")["valor"]
    somado = tidy[tidy["tipo"] == "componente"].groupby("ano_mes")["valor"].sum()
    comparacao = pd.concat([publicado, somado], axis=1, keys=["publicado", "somado"]).dropna()
    divergencia = (comparacao["somado"] - comparacao["publicado"]).abs()
    fora = comparacao[divergencia > comparacao["publicado"].abs() * 0.001]
    if len(fora):
        print(f"[tratamento] aviso: {len(fora)} meses em que os componentes não "
              f"fecham com o total publicado (ex.: {fora.index[0]}).")
    else:
        print(f"[tratamento] conferência OK: componentes = total geral em "
              f"{len(comparacao)} meses.")


def tratar(slug: str = SLUG) -> pd.DataFrame:
    """Transforma os XLSX brutos em DataFrame tidy e salva CSV. Retorna o DataFrame."""
    config.garantir_pastas()
    cfg = config.fonte(slug)
    pasta = _pasta_bruta(slug)
    metadados = json.loads((pasta / "_metadados.json").read_text(encoding="utf-8"))

    partes = [
        _ler_planilha(pasta / arquivo["nome"], cfg)
        for arquivo in metadados["arquivos"] if arquivo.get("tratar")
    ]
    if not partes:
        raise ValueError("Nenhum arquivo marcado para tratamento em _metadados.json.")

    tidy = _desambiguar(pd.concat(partes, ignore_index=True))
    tidy = _deflacionar(tidy, _indice_ipca(pasta))
    tidy = tidy.sort_values(["ano_mes", "tributo"]).reset_index(drop=True)
    _conferir_total(tidy, cfg)

    destino = config.DIR_PROCESSADOS / f"{slug}.csv"
    tidy.to_csv(destino, index=False, encoding="utf-8")
    print(f"[tratamento] {len(tidy)} linhas tidy salvas em {destino}")
    return tidy


if __name__ == "__main__":
    tratar()
